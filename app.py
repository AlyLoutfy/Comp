from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import io
import os
from werkzeug.utils import secure_filename
import tempfile
import json
from datetime import datetime

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Configure upload folder
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def compare_excel_sheets(modon_file, sakneen_file):
    """
    Compare Modon and Sakneen Excel sheets for status discrepancies.
    
    Logic:
    1. Use Unit Number (Modon) to match with UnitID (Sakneen)
    2. If EOI column in Modon has data, unit is not available
    3. If Sakneen Status is "Available" when it shouldn't be, highlight it
    """
    try:
        # Read Modon sheet (headers at row 3, index 2)
        modon_df = pd.read_excel(modon_file, header=2)
        
        # Read Sakneen sheet (headers at row 1, index 0)
        sakneen_df = pd.read_excel(sakneen_file, header=0)
        
        # Clean column names (remove extra spaces)
        modon_df.columns = modon_df.columns.str.strip()
        sakneen_df.columns = sakneen_df.columns.str.strip()
        
        print("Modon columns:", modon_df.columns.tolist())
        print("Sakneen columns:", sakneen_df.columns.tolist())
        
        # Find the correct column names (case-insensitive)
        modon_unit_col = None
        modon_eoi_col = None
        sakneen_unit_col = None
        sakneen_status_col = None
        
        for col in modon_df.columns:
            if 'unit number' in col.lower():
                modon_unit_col = col
            elif 'eoi' in col.lower():
                modon_eoi_col = col
        
        for col in sakneen_df.columns:
            if 'unitid' in col.lower() or 'unit id' in col.lower():
                sakneen_unit_col = col
            elif 'status' in col.lower():
                sakneen_status_col = col
        
        if not all([modon_unit_col, modon_eoi_col, sakneen_unit_col, sakneen_status_col]):
            missing_cols = []
            if not modon_unit_col:
                missing_cols.append("Unit Number in Modon")
            if not modon_eoi_col:
                missing_cols.append("EOI in Modon")
            if not sakneen_unit_col:
                missing_cols.append("UnitID in Sakneen")
            if not sakneen_status_col:
                missing_cols.append("Status in Sakneen")
            
            return {
                'error': f"Missing required columns: {', '.join(missing_cols)}",
                'modon_columns': modon_df.columns.tolist(),
                'sakneen_columns': sakneen_df.columns.tolist()
            }
        
        # Create a copy of Sakneen for highlighting
        sakneen_result = sakneen_df.copy()
        sakneen_result['Highlight'] = False
        sakneen_result['Issue'] = ''
        
        discrepancies = []
        
        # Create a mapping of Unit Number to EOI status from Modon
        modon_mapping = {}
        for idx, row in modon_df.iterrows():
            unit_number = row[modon_unit_col]
            eoi_value = row[modon_eoi_col]
            
            # Check if EOI has data (not null, not empty string, not NaN)
            has_eoi = pd.notna(eoi_value) and str(eoi_value).strip() != ''
            
            if pd.notna(unit_number):
                modon_mapping[str(unit_number).strip()] = has_eoi
        
        # Check each row in Sakneen
        for idx, row in sakneen_result.iterrows():
            unit_id = row[sakneen_unit_col]
            status = row[sakneen_status_col]
            
            if pd.notna(unit_id):
                unit_id_str = str(unit_id).strip()
                
                # Check if this unit exists in Modon
                if unit_id_str in modon_mapping:
                    has_eoi = modon_mapping[unit_id_str]
                    
                    # If Modon has EOI (unit not available) but Sakneen shows Available
                    if has_eoi and str(status).strip().lower() == 'available':
                        sakneen_result.at[idx, 'Highlight'] = True
                        sakneen_result.at[idx, 'Issue'] = f'Unit {unit_id} has EOI in Modon but shows as Available in Sakneen'
                        discrepancies.append({
                            'unit_id': unit_id,
                            'modon_status': 'Not Available (has EOI)',
                            'sakneen_status': status,
                            'issue': 'Unit has EOI in Modon but shows as Available in Sakneen'
                        })
        
        # Save the highlighted Sakneen file
        output_file = os.path.join(app.config['UPLOAD_FOLDER'], 'sakneen_highlighted.xlsx')
        
        # Create Excel writer with formatting
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            sakneen_result.to_excel(writer, sheet_name='Sakneen_With_Highlights', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Sakneen_With_Highlights']
            
            # Apply highlighting to rows that need attention
            from openpyxl.styles import PatternFill
            highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            
            for idx, row in sakneen_result.iterrows():
                if row['Highlight']:
                    # +2 because Excel is 1-indexed and we have a header row
                    excel_row = idx + 2
                    for col in range(1, len(sakneen_result.columns) + 1):
                        worksheet.cell(row=excel_row, column=col).fill = highlight_fill
        
        # Create export files for discrepancies
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Export discrepancies to CSV
        discrepancies_df = pd.DataFrame(discrepancies)
        csv_file = os.path.join(app.config['UPLOAD_FOLDER'], f'discrepancies_{timestamp}.csv')
        discrepancies_df.to_csv(csv_file, index=False)
        
        # Export discrepancies to Excel
        excel_discrepancies_file = os.path.join(app.config['UPLOAD_FOLDER'], f'discrepancies_{timestamp}.xlsx')
        with pd.ExcelWriter(excel_discrepancies_file, engine='openpyxl') as writer:
            discrepancies_df.to_excel(writer, sheet_name='Discrepancies', index=False)
        
        # Export discrepancies to JSON
        json_file = os.path.join(app.config['UPLOAD_FOLDER'], f'discrepancies_{timestamp}.json')
        with open(json_file, 'w') as f:
            json.dump({
                'timestamp': timestamp,
                'total_discrepancies': len(discrepancies),
                'discrepancies': discrepancies
            }, f, indent=2)
        
        return {
            'success': True,
            'discrepancies': discrepancies,
            'total_discrepancies': len(discrepancies),
            'output_file': output_file,
            'export_files': {
                'csv': csv_file,
                'excel': excel_discrepancies_file,
                'json': json_file
            },
            'modon_columns': modon_df.columns.tolist(),
            'sakneen_columns': sakneen_df.columns.tolist()
        }
        
    except Exception as e:
        return {'error': f'Error processing files: {str(e)}'}

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_files():
    if 'modon_file' not in request.files or 'sakneen_file' not in request.files:
        return jsonify({'error': 'Both Modon and Sakneen files are required'})
    
    modon_file = request.files['modon_file']
    sakneen_file = request.files['sakneen_file']
    
    if modon_file.filename == '' or sakneen_file.filename == '':
        return jsonify({'error': 'Both files must be selected'})
    
    if not (modon_file.filename.endswith('.xlsx') or modon_file.filename.endswith('.xls')):
        return jsonify({'error': 'Modon file must be an Excel file (.xlsx or .xls)'})
    
    if not (sakneen_file.filename.endswith('.xlsx') or sakneen_file.filename.endswith('.xls')):
        return jsonify({'error': 'Sakneen file must be an Excel file (.xlsx or .xls)'})
    
    try:
        # Save uploaded files
        modon_filename = secure_filename(modon_file.filename)
        sakneen_filename = secure_filename(sakneen_file.filename)
        
        modon_path = os.path.join(app.config['UPLOAD_FOLDER'], modon_filename)
        sakneen_path = os.path.join(app.config['UPLOAD_FOLDER'], sakneen_filename)
        
        modon_file.save(modon_path)
        sakneen_file.save(sakneen_path)
        
        # Compare the files
        result = compare_excel_sheets(modon_path, sakneen_path)
        
        # Clean up uploaded files
        os.remove(modon_path)
        os.remove(sakneen_path)
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'error': f'Error processing files: {str(e)}'})

@app.route('/download/<filename>')
def download_file(filename):
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        return jsonify({'error': 'File not found'}), 404

@app.route('/export/<export_type>/<filename>')
def export_file(export_type, filename):
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        return jsonify({'error': 'Export file not found'}), 404

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5002))
    app.run(debug=False, host='0.0.0.0', port=port)
